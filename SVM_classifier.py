"""
This code is to tune the SVM classifier to find its optimal hyper parameters.
"""

#Directory:
#C:\Users\pvtitor\Documents\1) Etudes\Tsinghua\M2\These\4) Python Code\Workspaces\Traitement Texte

#Load libraries
import openpyxl
import numpy
from matplotlib import pyplot
from pandas import read_csv
from pandas import set_option
from pandas.plotting import scatter_matrix
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.model_selection import KFold
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import classification_report
from sklearn.metrics import accuracy_score
from sklearn.pipeline import Pipeline
from sklearn.svm import SVC


DEBUG = False
def fprint(string):
    if DEBUG:
        print(string)


# Load dataset
url_dataset = "C:\\Users\\pvtitor\\Documents\\1) Etudes\\Tsinghua\\M2\\These\\7) Results\\3) Features\\CSV Results.csv"
url_features = "C:\\Users\\pvtitor\\Documents\\1) Etudes\\Tsinghua\\M2\\These\\7) Results\\3) Features\\CSV Results_features.csv"
print("Reading dataset")
dataset = read_csv(url_dataset, header=0, sep = ";")
features = read_csv(url_features, header=0, sep = ";")
print(dataset[0:10]), print()


"Analyse Data"
#shape
print(dataset.shape), print()
# types
set_option('display.max_rows', 5800)
print(dataset.dtypes), print()
# descriptions
set_option('precision', 3)
print(dataset.describe()), print()


"Visualize Data"

# histograms
features.hist(layout=(3,3), sharex=True, sharey=True, bins = 20)
pyplot.show(), print()

# density
features.plot(kind="Density", subplots=True, layout=(3,3), sharex=True, sharey=True, legend=True, fontsize=1)
pyplot.show(), print()


# correlation matrix
names = ['LCSQ', 'LCST', 'BoW', 'SM', 'WER', 'PER', 'POSt', 'WuPal']
fig = pyplot.figure()
ax = fig.add_subplot(111)
cax = ax.matshow(features.corr(), vmin=0, vmax=1, interpolation='none')
fig.colorbar(cax)
ticks = numpy.arange(0,8,1)
ax.set_xticks(ticks)
ax.set_yticks(ticks)
ax.set_xticklabels(names)
ax.set_yticklabels(names)
pyplot.show()



"Validation Dataset"
# Split-out validation dataset
array = dataset.values
X = array[:,0:7].astype(float)
print("X :"), print(X[0:10]), print()
Y = array[:,7]
print("Y :"), print(numpy.transpose(Y[0:10])), print()
validation_size = 0.20
seed = 7
X_train, X_validation, Y_train, Y_validation = train_test_split(X, Y, test_size=validation_size, random_state=seed)



"Tune Algorithm"

# Test options and evaluation metric
num_folds = 10
seed = 7
# Defining the scorings
scorings = ['accuracy', 'precision', 'recall']
score_accuracy = 'accuracy'
score_precsion = 'precision'
score_recall = 'recall'


# Tune scaled SVM
scaler = StandardScaler().fit(X_train)
rescaledX = scaler.transform(X_train)

c_values = [1.5]
kernel_values = ['rbf']
param_grid = dict(C=c_values, kernel=kernel_values)

model = SVC()
kfold = KFold(n_splits=num_folds, random_state=seed)


"Compute the grid search"

# Accuracy
print(), print("Accuracy")
grid_accuracy = GridSearchCV(estimator=model, param_grid=param_grid, scoring=score_accuracy, cv=kfold)
grid_result_accuracy = grid_accuracy.fit(rescaledX, Y_train)
print("Best accuracy: %f using %s" % (grid_result_accuracy.best_score_, grid_result_accuracy.best_params_)), print()

means_acc = grid_result_accuracy.cv_results_['mean_test_score']
print("Means"), print(means_acc), print()
stds_acc = grid_result_accuracy.cv_results_['std_test_score']
print("St Dev"), print(stds_acc), print()


# Precision
print(), print("Precision")
grid_precision = GridSearchCV(estimator=model, param_grid=param_grid, scoring=score_precsion, cv=kfold)
grid_result_precision = grid_precision.fit(rescaledX, Y_train)
print("Best precision: %f using %s" % (grid_result_precision.best_score_, grid_result_precision.best_params_)), print()

means_pre = grid_result_precision.cv_results_['mean_test_score']
print("Means"), print(means_pre), print()
stds_pre = grid_result_precision.cv_results_['std_test_score']
print("St Dev"), print(stds_pre), print()


# Recall
print(), print("Recall")
grid_recall = GridSearchCV(estimator=model, param_grid=param_grid, scoring=score_recall, cv=kfold)
grid_result_recall = grid_recall.fit(rescaledX, Y_train)
print("Best recal: %f using %s" % (grid_result_recall.best_score_, grid_result_recall.best_params_)), print()

means_rec = grid_result_recall.cv_results_['mean_test_score']
print("Means"), print(means_rec), print()
stds_rec = grid_result_recall.cv_results_['std_test_score']
print("St Dev"), print(stds_rec), print()

# params = grid_result_accuracy.cv_results_['params']
# for mean, stdev, param in zip(means, stds, params):
#    print("%f (%f) with: %r" % (mean, stdev, param))


# Creating tables with c values and kernel values for excel
excel_c_values = []
for c in c_values:
    for i in range(len(kernel_values)):
        excel_c_values.append(c)
print("C Values"), print(excel_c_values), print()

excel_kernel_values = []
for i in range(len(c_values)):
    for kernel in kernel_values:
        excel_kernel_values.append(kernel)
print("Kernel Values"), print(excel_kernel_values), print()



def model_to_xlsx(model_name, output_file, means_acc, means_pre, means_rec, c_values, kernel_values):

    #Setting up the Workbook Object
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name

    ws['A'+str(1)] = "Accuracy"
    ws['B'+str(1)] = "Precision"
    ws['C'+str(1)] = "Recall"
    ws['D'+str(1)] = "c"
    ws['E'+str(1)] = "kernel"

    list = []
    for i in range(len(means_acc)):
        list.append(i)

    for i, mean_acc, mean_pre, mean_rec, c, kernel in zip(list, means_acc, means_pre, means_rec, c_values, kernel_values):
        ws['A'+str(i+2)] = mean_acc
        ws['B'+str(i+2)] = mean_pre
        ws['C'+str(i+2)] = mean_rec
        ws['D'+str(i+2)] = c
        ws['E'+str(i+2)] = kernel


    wb.save(output_file)

model_name = "Cross Val SVM Tunning"
output_file = "Cross Val SVM Tunning.xlsx"
model_to_xlsx(model_name, output_file, means_acc, means_pre, means_rec, excel_c_values, excel_kernel_values)
