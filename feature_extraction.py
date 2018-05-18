"""
This code is to extract directly the vector of features from the pairs of sentences (train and test data).
"""

#Directory:
#C:\Users\pvtitor\Documents\1) Etudes\Tsinghua\M2\These\4) Python Code\Workspaces\Traitement Texte

# Import libraries
import numpy as np
import nltk
import openpyxl
import re

from nltk.corpus import wordnet as wn
from string import punctuation
from histogramme import plot_hist
from histogramme import plot_box
from itertools import product


# Function to print in case of need to spot the possible causes of mistake
DEBUG = False
def fprint(string):
    if DEBUG:
        print(string)


"Generic functions needed, for the implementation of every feature"

# TEXT NORMALIZATION
def strip_punctuation(string):
    return ''.join(c for c in string if c not in punctuation)

# Get tokens out of sentences
def get_tokens(sentence_1, sentence_2):
    sentence_no_punctuation_1 = strip_punctuation(sentence_1)
    tokens_1 = sentence_no_punctuation_1.strip().split(" ")
    sentence_no_punctuation_2 = strip_punctuation(sentence_2)
    tokens_2 = sentence_no_punctuation_2.strip().split(" ")
    return tokens_1, tokens_2

# Get the length of the shortest sentence to evaluate the similarity
def get_shorter_sentence(tokens_1, tokens_2):
    if len(tokens_1) <= len(tokens_2):
        return len(tokens_1), tokens_1, len(tokens_2), tokens_2
    else:
        return len(tokens_2), tokens_2, len(tokens_1), tokens_1

# Get the ground truth results to further make comparison
def get_true_results(data):
    results = []
    # Create a vector of the resuults of the true comparison
    for i in range(len(data)):
        a = data[i]
        # Extract the ith list of the train data array
        quality = int(a[0])
        # Extract the quality = result of the ith list
        results.append(quality)
        # Add this new result to the already existing list of results
    return results



"1) Bag of Words"

# CREATE A FUNCTION THAT WIILL ESTABLISH A DICTIONNARY FROM A PAIR OF SENTENCES
def get_dico(tokens_1, tokens_2):
    # We want to create directly a dictionnary out of a pari of sentences
    dico = {}
    word_count =  0
    for i in range(len(tokens_1)):
        word = tokens_1[i]
        # Store the words of sentence_1 in the dictionnary
        if not(word in dico.keys()):
            dico[word] = word_count
            word_count = word_count + 1
    for j in range(len(tokens_2)):
        word = tokens_2[j]
        # Store the words of sentence_2 in the dictionnary
        if not(word in dico.keys()):
            dico[word] = word_count
            word_count = word_count + 1
    return dico

# CREATE A FUNCTION THAT ASSOCIATE A BAG OF WORDS FROM A SENTENCE AND A DICTIONNARY
def get_bag_of_words(dico, tokens):
    BOW = np.zeros(len(dico))
    # We create this list and initialize it to zero for all its indexes.
    for word in tokens:
        index_dict = dico[word]
        # Extract the index of the vocabulary's dictionnary of the current word
        BOW[index_dict] += 1
        # Adds 1 to the value of this index, incrementing the multiplicity of this word in the bag of word
    return BOW

# DEFINE THE MEASURE OF SIMILARITY BETWEEN A PAIR OF BAG OF WORDS
def get_bow_ratio(BOW_1, BOW_2):
    # For now we will only consider the scalar product of the 2 bag of words to look at its similarity
    if len(BOW_1) != len(BOW_2):
        return "Error in lenghts of BOW"
        # Way of verifying that BOW_1 and BOW_2 have the same vocabulary dictionnary
    else:
        Scalar_Product = np.dot(BOW_1, BOW_2)
        norm_1 = np.linalg.norm(BOW_1)
        norm_2 = np.linalg.norm(BOW_2)
        measure = Scalar_Product / (norm_1*norm_2)
    return measure

# GET THE SIMILARITIES FOR EVERY PAIR OF SENTENCES
def get_bow_vector_of_ratios(data):
    vector_of_ratios = []
    for i in range(len(data)):

        sentence_1 = data[i][3]
        sentence_2 = data[i][4]
        tokens_1, tokens_2 = get_tokens(sentence_1, sentence_2)

        dico = get_dico(tokens_1, tokens_2)
        BOW_1 = get_bag_of_words(dico, tokens_1)
        BOW_2 = get_bag_of_words(dico, tokens_2)
        ratio = get_bow_ratio(BOW_1, BOW_2)
        # Check the result of the similarity between BOW_1 and BOW_2
        vector_of_ratios.append(round(ratio, 4))

    #plot_hist(vector_of_ratios),
    #plot_box(vector_of_ratios)
    return vector_of_ratios


"2) String Matching => weight can be changed"

# CREATE A FUNCTION THAT FINDS A MATCH OF A SUBSTRING OF length m IN A SENTENCE OF length n.
def get_stma_match(substring, sentence):
    m = len(substring)
    n = len(sentence)
    if m>n:
        fprint("Error in length between substring and sentence")
    else:
        for i in range(n-m):
            if substring == sentence[i:i+m]:
                fprint("substrings matched")
                fprint(substring)
                fprint(sentence[i:i+m])
                return 1, i
                # we need to return the index of the sentence_2 where the match starts as well so we do not take into account the words of the match for further matches.
        return 0, 0

# CREATE A FUNCTION THAT TRIES get_match FOR EVERY SUBSTRING OF length m FOR A PAIR OF SENTENCES
def try_every_stma_substring(tokens_1, tokens_2, m):
    # we assume that sentence_1 is shorter than sentence_2
    if len(tokens_1)<m:
        fprint("Error in the value of m")
        return 0, 0, 0
    elif len(tokens_1)>len(tokens_2):
        fprint("Error in the lengths of the two sentences")
        return 0, 0, 0
    else:
        for i in range(len(tokens_1)-m):
            substring = tokens_1[i:i+m]
            # we then try for this substring if it fits anywhere in sentence_2
            match, index_2 = get_stma_match(substring, tokens_2)
            if match: # it means that there has been a match for this substring
                return m, i, index_2
                # we need to return the index of the sentence_1 where the match starts as well to know which are the words who matched so we do not take them into account for further matches
        return 0, 0, 0

# CREATE THE FUNCTION THAT STORES THE lengthS OF MATCHES OF THE TWO SENTENCES, WE ASSUMNE THAT SENTENCE 1 IS SHORTER THAN SENTENCE 2
def get_vector_of_matches(tokens_1, tokens_2, vector_of_matches, l):
    n1, n2 = len(tokens_1), len(tokens_2)
    for i in range(l):
        match, index_1, index_2 = try_every_stma_substring(tokens_1, tokens_2, l-i)
        if match > 0: # it means that there has been a match of length n1-i of sentence_1 in sentence_2
            vector_of_matches.append(match)
            new_tokens_1 = tokens_1[0:index_1]+tokens_1[index_1+match:n1]
            new_tokens_2 = tokens_2[0:index_2]+tokens_2[index_2+match:n2]
            fprint("new_tokens_1, new_tokens_2")
            fprint(new_tokens_1)
            fprint(new_tokens_2)
            return get_vector_of_matches(new_tokens_1, new_tokens_2, vector_of_matches, l-i)
    return vector_of_matches

# DEFINE THE MEASURE OF SIMILARITY FOR THE StMa FEATURE
def get_stma_ratio(vector_of_matches, length, weight):
    # the dimension of "vector_of_matches" is the number of matches of substrings, and his coordinates contains the length of each match.
    # for simplification reasons, "vector_of_matches" has some 0 coordinates
    # the length is the number of words of the shortest sentence.
    # the weight is a parameter between 1 and 2 to be defined to find best results. The higher "weight" is, the most importance it is given to longer substrings.
    # the "threshold" is the value above which we assume the sentences to be paraphrase.
    ratio = 0
    for i in range(len(vector_of_matches)):
        ratio += vector_of_matches[i]**weight / length**weight
    return ratio

# GET THE SIMILARITIES FOR EVERY PAIR OF SENTENCES
def get_stma_vector_of_ratios(data):
    # Choose the weight
    weight = 1.2

    vector_of_ratios = []

    for i in range(len(data)):
        fprint(i)
        # Take out punctuation
        sequence_1 = strip_punctuation(data[i][3])
        sequence_2 = strip_punctuation(data[i][4])
        # Transform into vectors
        tokens_1, tokens_2 = get_tokens(sequence_1, sequence_2)
        if len(tokens_1)<=len(tokens_2):
            # Get the matches
            vector_of_matches = get_vector_of_matches(tokens_1, tokens_2, [], len(tokens_1))
            # Get the result of the string matching
            ratio = get_stma_ratio(vector_of_matches, len(tokens_1), weight)
        else:
            # Get the matches
            vector_of_matches = get_vector_of_matches(tokens_2, tokens_1, [], len(tokens_2))
            # Get the result of the string matching
            ratio = get_stma_ratio(vector_of_matches, len(tokens_2), weight)

        vector_of_ratios.append(round(ratio, 4))

    #plot_hist(vector_of_ratios)
    #plot_box(vector_of_ratios)
    return vector_of_ratios

"3) Longest Common Subsequence"

# CREAT A FUNCTION THAT RETURNS THE LCSQ BETWEEN A PAIR OF TOKENS USING A MATRIX
def get_lcsq(tokens_1, tokens_2):
    m = len(tokens_1)
    n = len(tokens_2)

    c = np.zeros_like(np.arange(m*n).reshape((m,n)))
    sequence = []

    # initialize c matrix
    for i in range(m):
        c[i,0] = 0
    for j in range(n):
        c[0,j] = 0

    # start filling c matrix
    for i in range(m):
        for j in range(n):
            if tokens_1[i] == tokens_2[j]:
                c[i,j] = c[i-1, j-1] + 1
                sequence.append(tokens_1[i])
            elif c[i-1, j] >= c[i, j-1]:
                c[i,j] = c[i-1, j]
            else:
                c[i,j] = c[i, j-1]

    fprint(c[m-1,n-1])
    fprint(sequence)
    return c[m-1,n-1], sequence

# DEFINE THE MEASURE OF SIMILARITY FOR THE LCSQ FEATURE
def get_lcsq_ratio(tokens_1, tokens_2):
    m, tok_1, n, tok_2 = get_shorter_sentence(tokens_1, tokens_2)
    length, sequence = get_lcsq(tok_1, tok_2)
    ratio = length / m
    return ratio

# GET THE SIMILARITIES FOR EVERY PAIR OF SENTENCES
def get_lcsq_vector_of_ratios(data):
    vector_of_ratios = []
    for i in range(len(data)):
        sentence_1 = data[i][3]
        sentence_2 = data[i][4]
        tokens_1, tokens_2 = get_tokens(sentence_1, sentence_2)
        ratio = get_lcsq_ratio(tokens_1, tokens_2)
        vector_of_ratios.append(round(ratio, 4))

    #plot_hist(vector_of_ratios)
    #plot_box(vector_of_ratios)
    return vector_of_ratios


"4) Longest Common Substring"

# TRY IF A SUBSTRING OF LENGTH P MATCHES IN ANY PART OF SENTENCE OF LENGTH N
def get_lcst_match(string, tokens):
    p = len(string)
    n = len(tokens)
    if p > n:
        fprint("Error in length between substring and tokens")
    else:
        for i in range(n-p):
            if string == tokens[i:i+p]:
                return 1
        return 0

# TRY IF ANY SUSBTRING OF LENGTH P MATCHES IN ANY PART OF SENTENCE OF LEGTH N
def try_every_substring(tokens_1, tokens_2, p):
    # we assume that sentence_1 is shorter than sentence_2
    m = len(tokens_1)
    n = len(tokens_2)
    if m < p:
        fprint("Error in the value of m")
        return 0
    elif m > n:
        fprint("Error in the lengths of the two sentences")
        return 0
    else:
        for i in range(m-p):
            string = tokens_1[i:i+p]
            if get_lcst_match(string, tokens_2):
                return p
        return 0

# DEFINE THE MEASURE OF SIMILARITY FOR THE LCST FEATURE
def get_lcst_ratio(tokens_1, tokens_2):
    m, tok_1, n, tok_2 = get_shorter_sentence(tokens_1, tokens_2)
    for i in range(m):
        p = try_every_substring(tok_1, tok_2, m-i)
        if p > 0:
            return p / m
    return 0

# GET THE SIMILARITIES FOR EVERY PAIR OF SENTENCES
def get_lcst_vector_of_ratios(data):
    vector_of_ratios = []
    for i in range(len(data)):
        sentence_1 = data[i][3]
        sentence_2 = data[i][4]
        tokens_1, tokens_2 = get_tokens(sentence_1, sentence_2)
        ratio = get_lcst_ratio(tokens_1, tokens_2)
        vector_of_ratios.append(round(ratio, 4))

    #plot_hist(vector_of_ratios)
    #plot_box(vector_of_ratios)
    return vector_of_ratios


"5) Word Error Rate"

# CREATE A FUNCTION THAT GET THE WER OF TWO LIST OF TOKENS USING A MATRIX
def get_wer(r, h):
    # creation of the d matrix
    d = np.zeros((len(r)+1)*(len(h)+1), dtype=np.uint8)
    d = d.reshape((len(r)+1, len(h)+1))

    # initialisation of the the d matrix
    for i in range(len(r)+1):
        for j in range(len(h)+1):
            if i == 0:
                d[0][j] = j
            elif j == 0:
                d[i][0] = i

    # computation of the the d matrix
    for i in range(1, len(r)+1):
        for j in range(1, len(h)+1):
            if r[i-1] == h[j-1]:
                d[i][j] = d[i-1][j-1]
            else:
                substitution = d[i-1][j-1] + 1
                insertion    = d[i][j-1] + 1
                deletion     = d[i-1][j] + 1
                d[i][j] = min(substitution, insertion, deletion)

    return d[len(r)][len(h)]

# DEFINE THE MEASURE OF SIMILARITY FOR THE WER FEATURE
def get_wer_ratio(tokens_1, tokens_2):
    m, tok_1, n, tok_2 = get_shorter_sentence(tokens_1, tokens_2)
    WER = get_wer(tok_1, tok_2)
    ratio = WER / n
    return ratio

# GET THE SIMILARITIES FOR EVERY PAIR OF SENTENCES
def get_wer_vector_of_ratios(data):
    vector_of_ratios = []
    for i in range(len(data)):
        sentence_1 = data[i][3]
        sentence_2 = data[i][4]
        tokens_1, tokens_2 = get_tokens(sentence_1, sentence_2)
        ratio = 1 - get_wer_ratio(tokens_1, tokens_2)
        ratio_norm = round(ratio, 4)
        vector_of_ratios.append(ratio_norm)

    #plot_hist(vector_of_ratios)
    #plot_box(vector_of_ratios)
    return vector_of_ratios


"6) Position Independent Word Error Rate"

# TO CREAT A BAG OF WORDS AND A DICO FROM A LIST OF TOKENS, WE USE THE FUNCTIONS OF THE BAG OF WORDS FEATURE

# DEFINE THE MEASURE OF SIMILARITY FOR THE WER FEATURE
def get_per_ratio(tokens_1, tokens_2):
    m, tok_1, n, tok_2 = get_shorter_sentence(tokens_1, tokens_2)
    l = abs(m-n)

    dico = get_dico(tok_1, tok_2)
    BOW_1 = get_bag_of_words(dico, tok_1)
    BOW_2 = get_bag_of_words(dico, tok_2)

    # Verify that BOW_1 and BOW_2 have the same vocabulary dictionnary
    if len(BOW_1) != len(BOW_2):
        return "Error in lenghts of BOW"
    else:
        c = 0
        for i in range(len(BOW_1)):
            c += abs(BOW_1[i] - BOW_2[i])
    score = 0.5 * (l + c)
    # The 0.5 is necessary because otherwise we would be counting twice every word appearing in 1 and not in 2 (and vice-versa).
    # And we would also be counting twice the substitutions (as word1 does not appear in 2 and word2 does not appear in 1).
    ratio = score / max(m, n)
    return ratio

# GET THE SIMILARITIES FOR EVERY PAIR OF SENTENCES
def get_per_vector_of_ratios(data):
    vector_of_ratios = []
    for i in range(len(data)):
        sentence_1 = data[i][3]
        sentence_2 = data[i][4]
        tokens_1, tokens_2 = get_tokens(sentence_1, sentence_2)
        ratio = 1 - get_per_ratio(tokens_1, tokens_2)
        ratio_norm = round(ratio, 4)
        vector_of_ratios.append(ratio_norm)

    #plot_hist(vector_of_ratios)
    #plot_box(vector_of_ratios)
    return vector_of_ratios


"7) Part-Of-Speech Comparison"

# GET A LIST OF TAGS FROM A LIST OF TOKENS
def get_tags(sentence):
    sentence_no_punctuation = strip_punctuation(sentence)
    tokens = sentence_no_punctuation.strip().split()
    words_and_tags = nltk.pos_tag(tokens)
    tags = []
    for i in range(len(words_and_tags)):
        tags.append(words_and_tags[i][1])
    return tags

# GET A DICO OF TAGS FROM TWO LIST OF TAGS
def get_dico_of_tags(tags_1, tags_2):
    dico = {}
    tag_count =  0

    for tag in tags_1:
        # Store the tags of tags_1 in the dictionnary
        if not(tag in dico.keys()):
            dico[tag] = tag_count
            tag_count = tag_count + 1

    for tag in tags_2:
        # Store the tags of tags_2 in the dictionnary
        if not(tag in dico.keys()):
            dico[tag] = tag_count
            tag_count += 1
    return dico

# GET A BAG OF TAGS FROM A DICO AND A LIST OF TAGS
def get_bag_of_tags(tags, dico):
    BOT = np.zeros(len(dico))
    for tag in tags:
        index_dict = dico[tag]
        BOT[index_dict] += 1
    return BOT

# DEFINE THE MEASURE OF SIMILARITY FOR THE WER FEATURE
def get_post_ratio(BOT_1, BOT_2):
    if len(BOT_1) != len(BOT_2):
        print("Error in the lenghts of vector of tags")
        return 0
    else:
        Scalar_Product = np.dot(BOT_1, BOT_2)
        norm_1 = np.linalg.norm(BOT_1)
        norm_2 = np.linalg.norm(BOT_2)
        ratio = Scalar_Product/(norm_1*norm_2)
        return ratio

# GET THE SIMILARITIES FOR EVERY PAIR OF SENTENCES
def get_post_vector_of_ratios(data):
    vector_of_ratios = []

    for i in range(len(data)):
        sentence_1 = data[i][3]
        sentence_2 = data[i][4]

        tags_1 = get_tags(sentence_1)
        tags_2 = get_tags(sentence_2)

        dico = get_dico_of_tags(tags_1, tags_2)

        BOT_1 = get_bag_of_tags(tags_1, dico)
        BOT_2 = get_bag_of_tags(tags_2, dico)

        ratio = get_post_ratio(BOT_1, BOT_2)
        vector_of_ratios.append(round(ratio, 4))

    #plot_hist(vector_of_ratios)
    #plot_box(vector_of_ratios)
    return vector_of_ratios


"8) Wu Palmer WordNet Similarity"

# GET ALL THE WORDS TAGGED AS NOUN, VERB, ADVERB OR ADJECTIVE FROM A LIST OF TOKENS
def get_wupal_list(sentence):
    list_of_tags = ["NOUN", "VERB", "ADJ", "ADV"]
    sentence_no_punctuation = strip_punctuation(sentence)
    tokens = sentence_no_punctuation.strip().split()
    words_and_tags = nltk.pos_tag(sentence, tagset='universal')
    list = []
    for i in range(len(words_and_tags)):
        if words_and_tags[i][1] in list_of_tags:
            list.append(words_and_tags[i][0])
    return list

# GET ALL SYNSETS FROM A WORD
def get_all_synsets(word):
    allsyn = set(ss for ss in wn.synsets(word))
    return allsyn

# COMPARE THE MEANING SIMILARITY OF TWO WORDS
def get_best_wup_similarity(word_1, word_2):
    allsyns1 = get_all_synsets(word_1)
    allsyns2 = get_all_synsets(word_2)
    #print("word_1", word_1, allsyns1)
    #print("word_2", word_2, allsyns2)
    if len(allsyns1) == 0:
        return 0, 0, 0
    elif len(allsyns2) ==0:
        return 0, 0, 0
    else:
        best_match = max((wn.wup_similarity(s1, s2) or 0, s1, s2) for s1, s2 in product(allsyns1, allsyns2))
        #print(best_match)
        return best_match

# FOR EACH WORD FROM LIST 1, STORE THE BEST SIMILARITY FROM THE COMPARISON OF EACH WORD FROM LIST 2
def get_wupal_ratio(list_1, list_2):
    scores = [0]*len(list_1)
    for i in range(len(list_1)):
        word_1 = list_1[i]
        score = 0
        for word_2 in list_2:
            result, s1, s2 = get_best_wup_similarity(word_1, word_2)
            if score < result:
                score = result
        scores[i] = score
    mean_score = sum(scores[i] for i in range(len(scores))) / len(list_1)
    return mean_score

# GET THE SIMILARITIES FOR EVERY PAIR OF SENTENCES
def get_wupal_vector_of_ratios(data):
    vector_of_ratios = []

    for i in range(len(data)):
        sentence_1 = data[i][3]
        sentence_2 = data[i][4]

        tokens_1, tokens_2 = get_tokens(sentence_1, sentence_2)
        if len(tokens_1) < len(tokens_2):
            list_1 = get_wupal_list(sentence_1)
            list_2 = get_wupal_list(sentence_2)
        else:
            list_1 = get_wupal_list(sentence_2)
            list_2 = get_wupal_list(sentence_1)

        ratio = get_wupal_ratio(list_1, list_2)
        vector_of_ratios.append(round(ratio, 4))

    #plot_hist(vector_of_ratios)
    #plot_box(vector_of_ratios)
    return vector_of_ratios



"Load Data"

# CREATE A FUNCTION THAT STORES TRAIN AND TEST DATA INTO VECTORS
def read_data():

    "Train Data"
    # Import in the object train_data the content of the train txt file (enter correct url)
    train_data_file = "C:\MSRParaphraseCorpus\msr_paraphrase_train.txt"
    # Create a file reader that will go from one line to the next at each step
    fr_train = open(train_data_file, "r", encoding="utf8")
    # Read the first line to avoid having the headers in the output
    line = fr_train.readline()

    # Create a null matrix that will contain all the information of the pairs of sentences of the train dataset readen from the file
    List_train = []
    for i in range(4072):
        # Skip to the next readline
        line = fr_train.readline()
        tokens = line.strip().split("\t")
        # Add the current line information to the matrix of train data elements
        List_train.append(tokens)
    fr_train.close()

    "Test Data"
    # Import in the object test_data the content of the test txt file (enter correct url)
    test_data_file = "C:\MSRParaphraseCorpus\msr_paraphrase_test.txt"
    # Create a file reader that will go from one line to the next at each step
    fr_test = open(test_data_file, "r", encoding="utf8")
    # Read the first line to avoid having the headers in the output
    line = fr_test.readline()

    # Create a null matrix that will contain all the information of the pairs of sentences of the test dataset readen from the file
    List_test = []
    for i in range(1725):
        # Skip to the next readline
        line = fr_test.readline()
        tokens = line.strip().split("\t")
        # Add the current line information to the matrix of train data elements
        List_test.append(tokens)
    fr_test.close()

    return List_train, List_test

# Load Train & Test Data
print("Reading train & test data...")
train_data, test_data = read_data()
print("done"), print()



"Get Results"

# Get Ground Truth Results
true_results_train = get_true_results(train_data)
true_results_test = get_true_results(test_data)

# Get ratios for all features
print("Gettinng Bag of Words similarities")
bow_ratios_train, bow_ratios_test = get_bow_vector_of_ratios(train_data), get_bow_vector_of_ratios(test_data)
print("done"), print()

print("Gettinng String Matching similarities")
stma_ratios_train, stma_ratios_test = get_stma_vector_of_ratios(train_data), get_stma_vector_of_ratios(test_data)
print("done"), print()

print("Gettinng Longest Commmon Subsequence similarities")
lcsq_ratios_train, lcsq_ratios_test = get_lcsq_vector_of_ratios(train_data), get_lcsq_vector_of_ratios(test_data)
print("done"), print()

print("Gettinng Longest Commmon Substring similarities")
lcst_ratios_train, lcst_ratios_test = get_lcst_vector_of_ratios(train_data), get_lcst_vector_of_ratios(test_data)
print("done"), print()

print("Gettinng Word Error Rate similarities")
wer_ratios_train, wer_ratios_test = get_wer_vector_of_ratios(train_data), get_wer_vector_of_ratios(test_data)
print("done"), print()

print("Gettinng Position Independent Word Error Rate similarities")
per_ratios_train, per_ratios_test = get_per_vector_of_ratios(train_data), get_per_vector_of_ratios(test_data)
print("done"), print()

print("Gettinng Part of Speech Tags similarities")
post_ratios_train, post_ratios_test = get_post_vector_of_ratios(train_data), get_post_vector_of_ratios(test_data)
print("done"), print()

#print("Gettinng Wu Palmer WordNet similarities")
#wupal_ratios_train, wupal_ratios_test = get_wupal_vector_of_ratios(train_data), get_wupal_vector_of_ratios(test_data)
#print("done"), print()



"Compile the results"

# Get the whol vector of ratios
print("Compiling the ratios into 1 single vector")
vector_of_ratios = [0]*18
vector_of_ratios[0], vector_of_ratios[9] = true_results_train, true_results_test
vector_of_ratios[1], vector_of_ratios[10] = bow_ratios_train, bow_ratios_test
vector_of_ratios[2], vector_of_ratios[11] = stma_ratios_train, stma_ratios_test
vector_of_ratios[3], vector_of_ratios[12] = lcsq_ratios_train, lcsq_ratios_test
vector_of_ratios[4], vector_of_ratios[13] = lcst_ratios_train, lcst_ratios_test
vector_of_ratios[5], vector_of_ratios[14] = wer_ratios_train, wer_ratios_test
vector_of_ratios[6], vector_of_ratios[15] = per_ratios_train, per_ratios_test
vector_of_ratios[7], vector_of_ratios[16] = post_ratios_train, post_ratios_test
#vector_of_ratios[8], vector_of_ratios[17] = wupal_ratios_train, wupal_ratios_test
print("done"), print()


"Save Data to Excel"

def model_to_xlsx(model_name, output_file, vector_of_ratios):
    N = len(vector_of_ratios[0])
    M = len(vector_of_ratios[9])

    # Setting up the Workbook Object
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name

    # First write the header
    print("Writing the header of excel file")
    ws['A'+str(1)] = "Ground Truth Train"
    ws['B'+str(1)] = "Ratio BoW Train"
    ws['C'+str(1)] = "Ratio StMa Train"
    ws['D'+str(1)] = "Ratio LCSQ Train"
    ws['E'+str(1)] = "Ratio LCST Train"
    ws['F'+str(1)] = "Ratio WER Train"
    ws['G'+str(1)] = "Ratio PER Train"
    ws['H'+str(1)] = "Ratio POSt Train"
    ws['I'+str(1)] = "Ratio WuPal Train"

    ws['K'+str(1)] = "Ground Truth Test"
    ws['L'+str(1)] = "Ratio BoW Test"
    ws['M'+str(1)] = "Ratio StMa Test"
    ws['N'+str(1)] = "Ratio LCSQ Test"
    ws['O'+str(1)] = "Ratio LCST Test"
    ws['P'+str(1)] = "Ratio WER Test"
    ws['Q'+str(1)] = "Ratio PER Test"
    ws['R'+str(1)] = "Ratio POSt Test"
    ws['S'+str(1)] = "Ratio WuPal Test"
    print("done"), print()

    print("Wrting the ratios in the excel file")
    for i in range(N):
        ws['A'+str(i+2)] = vector_of_ratios[0][i]
        ws['B'+str(i+2)] = vector_of_ratios[1][i]
        ws['C'+str(i+2)] = vector_of_ratios[2][i]
        ws['D'+str(i+2)] = vector_of_ratios[3][i]
        ws['E'+str(i+2)] = vector_of_ratios[4][i]
        ws['F'+str(i+2)] = vector_of_ratios[5][i]
        ws['G'+str(i+2)] = vector_of_ratios[6][i]
        ws['H'+str(i+2)] = vector_of_ratios[7][i]
        #ws['I'+str(i+2)] = vector_of_ratios[8][i]

    for i in range(M):
        ws['K'+str(i+2)] = vector_of_ratios[9][i]
        ws['L'+str(i+2)] = vector_of_ratios[10][i]
        ws['M'+str(i+2)] = vector_of_ratios[11][i]
        ws['N'+str(i+2)] = vector_of_ratios[12][i]
        ws['O'+str(i+2)] = vector_of_ratios[13][i]
        ws['P'+str(i+2)] = vector_of_ratios[14][i]
        ws['Q'+str(i+2)] = vector_of_ratios[15][i]
        ws['R'+str(i+2)] = vector_of_ratios[16][i]
        #ws['S'+str(i+2)] = vector_of_ratios[17][i]
    print("done"), print()

    wb.save(output_file)


def main():

    model_name = "Ratios for features"
    output_file = "Feature Extraction Ratios.xlsx"
    model_to_xlsx(model_name, output_file, vector_of_ratios)

if __name__ == '__main__':
    main()
